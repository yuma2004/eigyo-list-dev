"""
エクスポート API
"""
import io
import csv
from typing import Optional, List
from fastapi import APIRouter, HTTPException, Query, Response
from fastapi.responses import StreamingResponse
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment
import pandas as pd

from app.models.responses import ExportRequest
from app.services.company_service import CompanyService
from app.services.google_sheets import GoogleSheetsService

router = APIRouter()

# サービス初期化
google_sheets_service = GoogleSheetsService()
company_service = CompanyService(google_sheets_service)


@router.get("/csv")
async def export_csv(
    status: Optional[str] = Query(None, description="ステータスフィルター"),
    prefecture: Optional[str] = Query(None, description="都道府県フィルター"),
    industry: Optional[str] = Query(None, description="業界フィルター"),
    include_sales_status: bool = Query(True, description="営業ステータス含める")
):
    """
    CSV エクスポート
    
    Args:
        status: ステータスフィルター
        prefecture: 都道府県フィルター
        industry: 業界フィルター
        include_sales_status: 営業ステータス含める
        
    Returns:
        CSV ファイル
    """
    try:
        # フィルター条件構築
        filters = {}
        if status:
            filters["status"] = status
        if prefecture:
            filters["prefecture"] = prefecture
        if industry:
            filters["industry"] = industry
        
        # データ取得
        companies = await company_service.get_companies(filters=filters)
        
        # CSV データ作成
        output = io.StringIO()
        writer = csv.writer(output)
        
        # ヘッダー作成
        headers = [
            "ID", "会社名", "URL", "住所", "郵便番号", "都道府県", "市区町村",
            "住所詳細", "電話番号", "FAX番号", "代表者名", "事業内容",
            "設立年月日", "資本金", "問い合わせURL", "情報収集元URL",
            "作成日時", "更新日時"
        ]
        
        if include_sales_status:
            headers.extend([
                "営業ステータス", "メモ", "担当者", "最終コンタクト日", "次回アクション"
            ])
        
        writer.writerow(headers)
        
        # データ行作成
        for company in companies:
            row = [
                company.id,
                company.company_name,
                company.url,
                company.address,
                company.postal_code,
                company.prefecture,
                company.city,
                company.address_detail,
                company.tel,
                company.fax,
                company.representative,
                company.business_content,
                company.established_date,
                company.capital,
                company.contact_url,
                company.source_url,
                company.created_at.isoformat() if company.created_at else "",
                company.updated_at.isoformat() if company.updated_at else ""
            ]
            
            if include_sales_status and company.id:
                # 営業ステータス取得
                sales_status = await google_sheets_service.get_sales_status(company.id)
                if sales_status:
                    row.extend([
                        sales_status.status,
                        sales_status.memo,
                        sales_status.contact_person,
                        sales_status.last_contact_date.isoformat() if sales_status.last_contact_date else "",
                        sales_status.next_action
                    ])
                else:
                    row.extend(["", "", "", "", ""])
            
            writer.writerow(row)
        
        # レスポンス作成
        output.seek(0)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"companies_{timestamp}.csv"
        
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/excel")
async def export_excel(
    status: Optional[str] = Query(None, description="ステータスフィルター"),
    prefecture: Optional[str] = Query(None, description="都道府県フィルター"),
    industry: Optional[str] = Query(None, description="業界フィルター"),
    include_sales_status: bool = Query(True, description="営業ステータス含める")
):
    """
    Excel エクスポート
    
    Args:
        status: ステータスフィルター
        prefecture: 都道府県フィルター
        industry: 業界フィルター
        include_sales_status: 営業ステータス含める
        
    Returns:
        Excel ファイル
    """
    try:
        # フィルター条件構築
        filters = {}
        if status:
            filters["status"] = status
        if prefecture:
            filters["prefecture"] = prefecture
        if industry:
            filters["industry"] = industry
        
        # データ取得
        companies = await company_service.get_companies(filters=filters)
        
        # Excel ワークブック作成
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "企業リスト"
        
        # ヘッダースタイル
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center")
        
        # ヘッダー設定
        headers = [
            "ID", "会社名", "URL", "住所", "郵便番号", "都道府県", "市区町村",
            "住所詳細", "電話番号", "FAX番号", "代表者名", "事業内容",
            "設立年月日", "資本金", "問い合わせURL", "情報収集元URL",
            "作成日時", "更新日時"
        ]
        
        if include_sales_status:
            headers.extend([
                "営業ステータス", "メモ", "担当者", "最終コンタクト日", "次回アクション"
            ])
        
        # ヘッダー行作成
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
        
        # データ行作成
        for row_idx, company in enumerate(companies, 2):
            row_data = [
                company.id,
                company.company_name,
                company.url,
                company.address,
                company.postal_code,
                company.prefecture,
                company.city,
                company.address_detail,
                company.tel,
                company.fax,
                company.representative,
                company.business_content,
                company.established_date,
                company.capital,
                company.contact_url,
                company.source_url,
                company.created_at.isoformat() if company.created_at else "",
                company.updated_at.isoformat() if company.updated_at else ""
            ]
            
            if include_sales_status and company.id:
                # 営業ステータス取得
                sales_status = await google_sheets_service.get_sales_status(company.id)
                if sales_status:
                    row_data.extend([
                        sales_status.status,
                        sales_status.memo,
                        sales_status.contact_person,
                        sales_status.last_contact_date.isoformat() if sales_status.last_contact_date else "",
                        sales_status.next_action
                    ])
                else:
                    row_data.extend(["", "", "", "", ""])
            
            # セル設定
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col, value=value)
        
        # 列幅自動調整
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # ファイル出力
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"companies_{timestamp}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/template")
async def export_template():
    """
    インポート用テンプレート エクスポート
    
    Returns:
        テンプレートExcelファイル
    """
    try:
        # テンプレート作成
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "企業情報テンプレート"
        
        # ヘッダースタイル
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center")
        
        # ヘッダー設定
        headers = [
            "会社名*", "URL*", "住所", "電話番号", "FAX番号", "代表者名",
            "事業内容", "設立年月日", "資本金", "問い合わせURL"
        ]
        
        # ヘッダー行作成（背景色付き）
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = openpyxl.styles.PatternFill(
                start_color="4472C4",
                end_color="4472C4",
                fill_type="solid"
            )
        
        # サンプルデータ行
        sample_data = [
            "サンプル株式会社",
            "https://example.com",
            "東京都渋谷区1-1-1",
            "03-1234-5678",
            "03-1234-5679",
            "山田太郎",
            "ITサービス業",
            "2020-01-01",
            "1000万円",
            "https://example.com/contact"
        ]
        
        for col, value in enumerate(sample_data, 1):
            ws.cell(row=2, column=col, value=value)
        
        # 注意事項シート追加
        notes_ws = wb.create_sheet("注意事項")
        notes = [
            "【企業情報インポート用テンプレート】",
            "",
            "■ 必須項目",
            "・会社名: 必須入力",
            "・URL: 必須入力（https://で始まる有効なURL）",
            "",
            "■ 入力規則",
            "・電話番号: ハイフン区切りで入力（例: 03-1234-5678）",
            "・設立年月日: YYYY-MM-DD形式で入力（例: 2020-01-01）",
            "・URL: https://またはhttp://で始まる有効なURL",
            "",
            "■ 使用方法",
            "1. サンプル行を参考にデータを入力",
            "2. 1行目のヘッダーは変更しないでください",
            "3. 入力完了後、APIのインポート機能を使用してアップロード"
        ]
        
        for row, note in enumerate(notes, 1):
            notes_ws.cell(row=row, column=1, value=note)
        
        # 列幅調整
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 注意事項シートの列幅調整
        notes_ws.column_dimensions['A'].width = 80
        
        # ファイル出力
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=companies_template.xlsx"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/stats")
async def get_export_stats():
    """
    エクスポート統計情報取得
    
    Returns:
        統計情報
    """
    try:
        # 企業数取得
        companies = await company_service.get_companies()
        total_companies = len(companies)
        
        # ステータス別集計
        status_summary = await google_sheets_service.get_sales_summary()
        
        # 都道府県別集計
        prefecture_summary = {}
        for company in companies:
            if company.prefecture:
                prefecture_summary[company.prefecture] = prefecture_summary.get(company.prefecture, 0) + 1
        
        return {
            "success": True,
            "total_companies": total_companies,
            "status_summary": status_summary,
            "prefecture_summary": prefecture_summary,
            "last_updated": datetime.now().isoformat(),
            "message": "Export statistics retrieved successfully"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))